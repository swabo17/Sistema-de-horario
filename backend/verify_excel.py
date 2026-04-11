import openpyxl
from pathlib import Path

# Archivo más reciente
xlsx_file = Path("generated_schedules/horario_Sucursal_Centro_20260411_003213.xlsx")

if xlsx_file.exists():
    print(f"✓ Archivo encontrado: {xlsx_file}")
    
    # Cargar workbook
    wb = openpyxl.load_workbook(xlsx_file)
    ws = wb.active
    
    print(f"\n📄 Hoja activa: {ws.title}")
    print(f"📐 Dimensiones: {ws.dimensions}")
    
    # Revisar primeras 5 celdas
    print(f"\n📌 Encabezados:")
    print(f"A1 (Título): {ws['A1'].value}")
    print(f"A2 (Empleado): {ws['A2'].value}")
    print(f"B2 (Día 1): {ws['B2'].value}")
    print(f"A3 (Día semana): {ws['A3'].value}")
    print(f"B3 (Primer día): {ws['B3'].value}")
    
    # Verificar datos
    print(f"\n👤 Empleados (filas 4+):")
    for row in range(4, 7):
        emp_name = ws[f'A{row}'].value
        if emp_name:
            print(f"  Fila {row}: {emp_name}")
            # Ver primeros dias
            for col in range(2, 7):
                cell_value = ws.cell(row, col).value
                print(f"    Día {col-1}: {cell_value}", end=" ")
            print()
    
    print(f"\n✅ Estructura del Excel VÁLIDA")
else:
    print(f"❌ Archivo no encontrado: {xlsx_file}")
